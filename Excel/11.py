from clearConsole import *
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


clearConsole()

wb = load_workbook("regions.xlsx")
ws = wb.active 
df = pd.read_excel("all_shifts.xlsx")

df_1 = df[["Sales Rep", "Cost per","Units Sold"]]
df_1["Total"] = df_1["Cost per"] * df_1["Units Sold"]

# print(df_1)

rows = dataframe_to_rows(df_1, index=False)

for  r_index, row in enumerate(rows,1):
    for c_index , col in enumerate(row, 1):
        ws.cell(row = r_index, column = c_index, value = col)


wb.save("combined.xlsx")

print("DONE")