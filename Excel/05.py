# OpenPyxl
import pandas as pd
from openpyxl.workbook import workbook
from openpyxl import Workbook, load_workbook
import os

def clearConsole():
    command = 'clear'
    if os.name in ('nt', 'dos'):  # If Machine is running on Windows, use cls
        command = 'cls'
    os.system(command)

clearConsole()


wb = Workbook()

ws = wb.active

ws1 =wb.create_sheet("NewSheet")
ws2 =wb.create_sheet("Another", 0)

ws.title = "MySheet"

wb2= load_workbook("regions.xlsx")

new_sheet = wb2.create_sheet("NewSheet")
active_sheet = wb2.active


cell = active_sheet["A1"]

print(cell.value)


active_sheet["A1"] = 0 
wb2.save("modified.xlsx")

s2= pd.read_excel("modified.xlsx")
print(s2)

print(wb.sheetnames)


print("IKo")